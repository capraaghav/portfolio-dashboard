"""Broker CSV/Excel parsing and normalisation into a single schema.

Output schema (one row per holding-per-account):
    ticker, shares, avg_cost, ltp, current_value, pnl, account,
    isin, sector, qty_long_term, purchase_date
"""

from __future__ import annotations
import io
import re
import openpyxl
import pandas as pd

# ─── Broker format definitions ────────────────────────────────────────────────

BROKER_FORMATS = {
    "zerodha_kite": {
        "required": ["Instrument", "Qty."],
        "map": {
            "Instrument": "ticker",
            "Qty.": "shares",
            "Avg. cost": "avg_cost",
            "LTP": "ltp",
            "Cur. val": "current_value",
            "P&L": "pnl",
        },
    },
    "zerodha_console": {
        "required": ["Symbol", "ISIN", "Quantity", "Average Price"],
        "map": {
            "Symbol": "ticker",
            "ISIN": "isin",
            "Quantity": "shares",
            "Average Price": "avg_cost",
            "Last Price": "ltp",
            "Current Value": "current_value",
            "P&L": "pnl",
        },
    },
    # Zerodha Console Excel/detailed export (has Sector + Quantity Available + Long Term)
    "zerodha_console_v2": {
        "required": ["Symbol", "ISIN", "Sector", "Quantity Available", "Average Price"],
        "map": {
            "Symbol": "ticker",
            "ISIN": "isin",
            "Sector": "sector",
            "Quantity Available": "shares",
            "Quantity Long Term": "qty_long_term",
            "Average Price": "avg_cost",
            "Previous Closing Price": "ltp",
            "Unrealized P&L": "pnl",
        },
    },
    "groww": {
        "required": ["Stock Symbol", "Quantity", "Average Price"],
        "map": {
            "Stock Symbol": "ticker",
            "Quantity": "shares",
            "Average Price": "avg_cost",
            "Current Market Price": "ltp",
            "Current Value": "current_value",
            "Total Returns": "pnl",
        },
    },
    "upstox": {
        "required": ["Sym", "Qty", "Avg. Price"],
        "map": {
            "Sym": "ticker",
            "Qty": "shares",
            "Avg. Price": "avg_cost",
            "LTP": "ltp",
            "Curr. Val.": "current_value",
            "P&L": "pnl",
        },
    },
    "angel": {
        "required": ["NSE Symbol", "Quantity", "Avg. Buy Price"],
        "map": {
            "NSE Symbol": "ticker",
            "Quantity": "shares",
            "Avg. Buy Price": "avg_cost",
            "Current Price": "ltp",
            "Current Value": "current_value",
            "Unrealised P&L": "pnl",
        },
    },
    # HDFC Securities holdings — Symbol is the full company NAME (needs resolution),
    # and there is no cost basis (only LTP + value + quantity).
    "hdfc": {
        "required": ["Symbol", "LTP", "STOCK VAL.", "AVAIL QTY"],
        "map": {
            "Symbol": "ticker",
            "LTP": "ltp",
            "STOCK VAL.": "current_value",
            "AVAIL QTY": "shares",
        },
    },
}

# Column aliases for unknown broker CSVs
GENERIC_ALIASES = {
    "ticker":        ["ticker", "symbol", "stock", "scrip", "instrument", "code", "stock symbol",
                      "trading symbol", "sym", "nse symbol", "bse symbol", "script"],
    "shares":        ["shares", "qty", "qty.", "quantity", "quantity available", "units",
                      "holding qty", "no. of shares", "no of shares", "net qty",
                      "avail qty", "available qty", "free qty", "net quantity"],
    "avg_cost":      ["avg cost", "avg. cost", "average cost", "average price", "buy price",
                      "cost price", "avg price", "avg. price", "avg buy price", "avg. buy price",
                      "purchase price"],
    "ltp":           ["ltp", "last price", "current price", "cmp", "market price", "price",
                      "last traded price", "current market price"],
    "current_value": ["current value", "cur. val", "cur. val.", "market value", "present value",
                      "curr. val.", "current market value", "stock val.", "stock val", "stock value",
                      "value"],
    "pnl":           ["p&l", "pnl", "profit/loss", "gain/loss", "returns", "total returns",
                      "unrealized p&l", "unrealised p&l", "net p&l"],
    "isin":          ["isin", "isin code", "isin_code"],
    "sector":        ["sector", "industry"],
    "qty_long_term": ["quantity long term", "long term qty", "lt quantity"],
    "purchase_date": ["purchase date", "buy date", "trade date", "date", "order execution time",
                      "transaction date"],
}

# Rows to drop — summary lines appended by some brokers
_SKIP_TICKERS = {"nan", "none", "null", "total", "grand total", "instrument", "net total", ""}

# Substrings that signal a header row (matched against underscore-normalised cells)
_HEADER_KEYWORDS = {
    "symbol", "instrument", "ticker", "scrip", "isin",
    "qty", "quantity", "shares", "ltp",
}

_NUMERIC_COLS = ["shares", "avg_cost", "ltp", "current_value", "pnl", "qty_long_term"]
_KEEP_COLS = ["ticker", "shares", "avg_cost", "ltp", "current_value", "pnl",
              "account", "isin", "sector", "qty_long_term", "purchase_date"]


# ─── Company-name handling (brokers like HDFC put the full name in "Symbol") ──

_SYMBOL_RE = re.compile(r"^[A-Z0-9&.\-]+$")
_LIMITED_RE = re.compile(r"\b(LIMITED|LTD)\b")
# Broker instrument junk after the company name: '-EQ', '-EQ1/-', ' - EQ 2', 'RE1',
# 'EQUITY', NSE series codes, face values. Cut at the first such marker.
_TRAILING_JUNK_RE = re.compile(
    r"[\s\-]+(EQUITY|EQ|BE|BZ|BL|SM|RT|GS|GB|NEW|FV|F\.V\.|SR|NCD|BOND|RS|RE)(?![A-Z])")
# NSE series suffixes appended to clean tickers (INDOWIND-BE, DIL-BZ) — not part of
# the Yahoo symbol. (Won't touch real hyphenated symbols like BAJAJ-AUTO.)
_SERIES_SUFFIX_RE = re.compile(r"-(BE|BZ|BL|SM|RT|EQ|GS|GB|IL|IQ|N[12])$")
_ISIN_RE = re.compile(r"^IN[A-Z][0-9A-Z]{9}$")


def is_isin(s: str) -> bool:
    """True for an Indian ISIN like INE002A01018 (12 chars, starts with IN)."""
    return bool(_ISIN_RE.match(str(s).strip().upper()))


def looks_like_symbol(s: str) -> bool:
    """True if the value is already a clean ticker (e.g. RELIANCE, BAJAJ-AUTO),
    not a verbose company name (e.g. 'ABB INDIA LIMITED EQ NEW RS. 2/-')."""
    s = str(s).strip().upper()
    return " " not in s and 1 <= len(s) <= 15 and bool(_SYMBOL_RE.match(s))


def classify_asset(name: str) -> str:
    """Equity / Mutual Fund / Bond-NCD — so we don't try to price bonds & funds as stocks."""
    u = str(name).upper()
    if "MUTUAL FUND" in u or "MUTUALFUND" in u or " MF " in f" {u} ":
        return "Mutual Fund"
    if "NCD" in u or "DEBENTURE" in u or "BOND" in u or re.search(r"FVRS\s*\d", u):
        return "Bond/NCD"
    return "Equity"


def clean_company_name(name: str) -> str:
    """Strip broker instrument suffixes to get a searchable company name.
    'ABB INDIA LIMITED EQ NEW RS. 2/-' -> 'ABB INDIA LIMITED'
    'BHARAT ELECTRONIC-EQ'            -> 'BHARAT ELECTRONIC'
    'ADANI WILMAR-EQ1/-'             -> 'ADANI WILMAR'
    'HINDUSTAN UNILEV RE1'           -> 'HINDUSTAN UNILEV'"""
    s = str(name).upper().replace("''", "").replace('"', "").strip()
    m = _LIMITED_RE.search(s)
    if m:
        s = s[:m.end()]                       # keep up to and including LIMITED/LTD
    else:
        cut = _TRAILING_JUNK_RE.search(s)
        if cut:
            s = s[:cut.start()]
    s = re.sub(r"[\s\-/]+$", "", s)            # trailing dashes/slashes/space
    return re.sub(r"\s+", " ", s).strip()


# ─── Helpers ──────────────────────────────────────────────────────────────────

def detect_broker(df: pd.DataFrame) -> str:
    for broker, info in BROKER_FORMATS.items():
        if all(col in df.columns for col in info["required"]):
            return broker
    return "generic"


def normalize_generic(df: pd.DataFrame) -> pd.DataFrame:
    # Normalise underscores → spaces so UPPER_SNAKE_CASE headers (NSE_SYMBOL, COST_PRICE,
    # ISIN_CODE) match the space-separated aliases.
    col_lower = {c.lower().strip().replace("_", " "): c for c in df.columns}
    mapping = {}
    for field, aliases in GENERIC_ALIASES.items():
        for alias in aliases:
            key = alias.replace("_", " ")
            if key in col_lower and col_lower[key] not in mapping.values():
                mapping[col_lower[key]] = field
                break
    return df.rename(columns=mapping)


def _header_score(cells: list[str]) -> int:
    """How header-like a row is — count cells containing a header keyword."""
    return sum(1 for c in cells for kw in _HEADER_KEYWORDS if kw in c)


def find_excel_header(raw_bytes: bytes) -> tuple:
    """Return (sheet_name, header_row_index) for the best data header across all sheets.
    Handles multi-sheet workbooks and UPPER_SNAKE_CASE headers buried under metadata rows.
    Must NOT use read_only=True — that shifts row indices vs pd.read_excel(skiprows=N)."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    sheets = wb.sheetnames
    best = (sheets[0], 0, 0)  # sheet, row, score
    for sheet in sheets:
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(v).lower().strip().replace("_", " ") for v in row if v is not None]
            if len(cells) < 3:
                continue
            score = _header_score(cells)
            if score > best[2]:
                best = (sheet, i, score)
            if i > 40:
                break
    wb.close()
    return (best[0], best[1]) if best[2] >= 2 else (sheets[0], 0)


def find_csv_header_row(lines: list[str]) -> int:
    """Return the line index that looks most like a CSV header (most commas, mostly non-numeric)."""
    best_idx, best_score = 0, -1
    for i, line in enumerate(lines[:30]):
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue
        non_num = sum(1 for p in parts if p and not p.replace(".", "").replace("-", "").isnumeric())
        score = len(parts) + non_num * 2
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


# ─── PDF statements (e.g. IIFL Portfolio+) ───────────────────────────────────

_PDF_TICKER_RE = re.compile(r"^[A-Z][A-Z0-9&.\-]{0,18}$")


def _pdf_num(tok: str):
    """Parse one PDF numeric token: Indian commas (1,09,912.73), parenthesised
    negatives ((-7,400.64)), trailing %, and '--' blanks."""
    t = tok.replace(",", "").replace("%", "").strip()
    neg = t.startswith("(")
    t = t.strip("()")
    if t in ("--", "-", "", "nan", "na"):
        return None
    try:
        v = float(t)
        return -abs(v) if neg else v
    except ValueError:
        return None


def _parse_pdf_line(line: str) -> dict | None:
    """Pull one holding from a PDF text line 'SCRIP qty buyprice invested ltp curval …'.
    Sector headings, column headers and footers are skipped — they aren't an
    ALL-CAPS ticker followed by numbers."""
    parts = line.split()
    if len(parts) < 6:
        return None
    name = parts[0].upper()
    if not _PDF_TICKER_RE.match(name):
        return None
    nums = [_pdf_num(t) for t in parts[1:]]
    if sum(1 for x in nums if x is not None) < 5:
        return None
    # Positional columns: Qty, Purchase Price, Invested Value, Current Price, Current Value
    if nums[0] is None or nums[1] is None or nums[4] is None:
        return None
    if not (nums[0] > 0 and nums[1] > 0):
        return None
    return {"ticker": name, "shares": nums[0], "avg_cost": nums[1],
            "ltp": nums[3], "current_value": nums[4]}


def parse_pdf(raw_bytes: bytes) -> pd.DataFrame:
    """Extract holdings from a portfolio-statement PDF (e.g. IIFL Portfolio+ report).
    Reads every page's text and keeps lines shaped like 'TICKER qty price …'."""
    try:
        import pdfplumber
    except ImportError:
        raise ValueError("PDF support needs the 'pdfplumber' package — run: pip install pdfplumber")

    rows = []
    with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                row = _parse_pdf_line(line)
                if row:
                    rows.append(row)
    if not rows:
        raise ValueError(
            "Couldn't find a holdings table in this PDF — it may be scanned (image-only) "
            "or use an unusual layout. Try a CSV/Excel export instead."
        )
    return pd.DataFrame(rows)


def parse_uploaded_file(f, account_name: str) -> pd.DataFrame:
    """Parse one uploaded file-like object into the normalised schema."""
    name = f.name.lower()
    is_pdf = name.endswith(".pdf")

    if is_pdf:
        df = parse_pdf(f.read())                       # already normalised schema
    elif name.endswith((".xlsx", ".xls")):
        raw_bytes = f.read()
        sheet, header_row = find_excel_header(raw_bytes)
        df = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=sheet, skiprows=header_row)
    else:
        content = f.read().decode("utf-8", errors="replace")
        lines = [l for l in content.split("\n") if l.strip()]
        header_idx = find_csv_header_row(lines)
        df = pd.read_csv(io.StringIO("\n".join(lines[header_idx:])))

    if df is None or df.empty:
        raise ValueError("Could not read any data from file.")

    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]

    if is_pdf:
        pass                                           # parse_pdf already normalised columns
    elif detect_broker(df) != "generic":
        df = df.rename(columns=BROKER_FORMATS[detect_broker(df)]["map"])
    else:
        df = normalize_generic(df)

    df["account"] = account_name
    df = df[[c for c in _KEEP_COLS if c in df.columns]].copy()

    # Coalesce: when the symbol is missing/null but an ISIN exists, use the ISIN
    # (it gets resolved to a ticker downstream). Supports depository-style exports.
    _BLANK = {"", "null", "nan", "none"}
    if "isin" in df.columns:
        if "ticker" not in df.columns:
            df["ticker"] = pd.NA
        blank = df["ticker"].isna() | df["ticker"].astype(str).str.strip().str.lower().isin(_BLANK)
        df.loc[blank, "ticker"] = df.loc[blank, "isin"]

    if "ticker" not in df.columns:
        raise ValueError(
            "Could not find a ticker/symbol or ISIN column. Check that the file has a column "
            "named Symbol, Instrument, Stock Symbol, NSE_SYMBOL, ISIN, etc."
        )

    df["ticker"] = (
        df["ticker"].astype(str).str.strip()
        .str.replace(r"\.(NS|BO|BSE|NSE)$", "", regex=True)
        .str.upper()
        .str.replace(_SERIES_SUFFIX_RE, "", regex=True)   # drop -BE/-BZ/… NSE series codes
    )
    # Keep real tickers/names only — drops footer/disclaimer paragraphs that some
    # broker exports append as a trailing row (no real symbol or name is this long).
    df = df[df["ticker"].str.len().between(1, 50)]
    df = df[~df["ticker"].str.lower().isin(_SKIP_TICKERS)]

    for col in _NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(r"[₹,\s%]", "", regex=True),
                errors="coerce",
            )

    if "purchase_date" in df.columns:
        df["purchase_date"] = pd.to_datetime(df["purchase_date"], errors="coerce", dayfirst=True)

    return df.reset_index(drop=True)


def parse_all(uploaded, account_names: dict) -> tuple[pd.DataFrame | None, list]:
    """Parse a list of uploaded files. Returns (combined_df_or_None, errors)."""
    all_dfs, errors = [], []
    for f in uploaded:
        label = account_names.get(f.name, f.name)
        try:
            df = parse_uploaded_file(f, label)
            if len(df) == 0:
                errors.append((f.name, "No holdings rows found after parsing."))
            else:
                all_dfs.append(df)
        except Exception as e:  # noqa: BLE001 — surface any parse failure to the UI
            errors.append((f.name, str(e)))

    if not all_dfs:
        return None, errors
    return pd.concat(all_dfs, ignore_index=True), errors
